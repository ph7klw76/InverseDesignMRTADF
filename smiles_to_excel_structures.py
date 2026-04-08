#!/usr/bin/env python3
"""
Insert chemical structure images into an Excel sheet from a SMILES column.

Supports:
- .xlsx input
- .csv input (converted to .xlsx first)

Examples:
    python smiles_to_excel_structures.py ^
        --input top50_candidates-2.csv ^
        --output top50_candidates_with_structures-2.xlsx ^
        --smiles-column smiles ^
        --image-column B

    python smiles_to_excel_structures.py ^
        --input molecules.xlsx ^
        --output molecules_with_structures.xlsx ^
        --smiles-column A ^
        --image-column B
"""

import argparse
import os
import shutil
import string
import sys
import tempfile
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from rdkit import Chem
from rdkit.Chem import AllChem, Draw


def excel_col_to_index(col):
    """
    Convert Excel column label like A, B, AA to 1-based index.
    Accepts integers too.
    """
    if isinstance(col, int):
        if col < 1:
            raise ValueError(f"Invalid column index {col}")
        return col

    text = str(col).strip()

    if text.isdigit():
        value = int(text)
        if value < 1:
            raise ValueError(f"Invalid column index {value}")
        return value

    # Only allow true Excel-style letters here
    if not text.isalpha():
        raise ValueError(f"Invalid Excel column label: {text}")

    text = text.upper()
    value = 0
    for ch in text:
        if ch not in string.ascii_uppercase:
            raise ValueError(f"Invalid Excel column label: {text}")
        value = value * 26 + (ord(ch) - ord("A") + 1)

    if value < 1:
        raise ValueError(f"Invalid column index {value}")
    return value


def index_to_excel_col(idx):
    """Convert 1-based index to Excel column letters."""
    if idx < 1:
        raise ValueError(f"Invalid column index {idx}")

    letters = []
    while idx:
        idx, rem = divmod(idx - 1, 26)
        letters.append(chr(rem + ord("A")))
    return "".join(reversed(letters))


def resolve_column(ws, col_spec, header_row=1):
    """
    Resolve a column spec from:
    - integer index
    - numeric string like '2'
    - header name like 'smiles'
    - Excel letters like 'A', 'B', 'AA'
    """

    if isinstance(col_spec, int):
        return excel_col_to_index(col_spec)

    text = str(col_spec).strip()

    # 1. numeric index
    if text.isdigit():
        return excel_col_to_index(text)

    # 2. header name exact match first
    for col_idx in range(1, ws.max_column + 1):
        cell_value = ws.cell(row=header_row, column=col_idx).value
        if cell_value is not None and str(cell_value).strip() == text:
            return col_idx

    # 3. header name case-insensitive fallback
    lower_text = text.lower()
    for col_idx in range(1, ws.max_column + 1):
        cell_value = ws.cell(row=header_row, column=col_idx).value
        if cell_value is not None and str(cell_value).strip().lower() == lower_text:
            return col_idx

    # 4. Excel letters only if it looks like a real Excel column label.
    # Limit length to avoid interpreting words like "smiles" as Excel columns.
    if text.isalpha() and len(text) <= 3:
        return excel_col_to_index(text)

    headers = [
        str(ws.cell(row=header_row, column=i).value).strip()
        for i in range(1, ws.max_column + 1)
        if ws.cell(row=header_row, column=i).value is not None
    ]
    raise ValueError(
        f"Could not resolve column '{col_spec}'. "
        f"Available headers in row {header_row}: {headers}"
    )


def prepare_workbook_from_input(input_path, output_path, sheet_name=None):
    """
    If input is CSV, convert to XLSX.
    If input is XLSX, copy to output path.
    """
    input_path = Path(input_path)
    output_path = Path(output_path)

    if input_path.suffix.lower() == ".csv":
        df = pd.read_csv(input_path)
        with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
            df.to_excel(writer, index=False, sheet_name=sheet_name or "Sheet1")
        return output_path

    if input_path.suffix.lower() == ".xlsx":
        shutil.copyfile(input_path, output_path)
        return output_path

    raise ValueError("Input file must be .csv or .xlsx")


def smiles_to_mol(smiles):
    """Convert SMILES to RDKit Mol."""
    if smiles is None:
        return None

    smiles = str(smiles).strip()
    if not smiles:
        return None

    mol = Chem.MolFromSmiles(smiles)
    if mol is None:
        return None

    try:
        AllChem.Compute2DCoords(mol)
    except Exception:
        pass

    return mol


def make_structure_png(mol, png_path, img_width=250, img_height=180):
    """Draw molecule as PNG."""
    img = Draw.MolToImage(mol, size=(img_width, img_height))
    img.save(png_path)


def autoset_column_width(ws, column_index, width):
    col_letter = index_to_excel_col(column_index)
    ws.column_dimensions[col_letter].width = width


def process_workbook(
    workbook_path,
    sheet_name,
    smiles_column,
    image_column,
    header_row=1,
    start_row=None,
    image_header="Structure",
    img_width=250,
    img_height=180,
    row_height=140,
):
    wb = load_workbook(workbook_path)
    ws = wb[sheet_name] if sheet_name else wb.active

    smiles_col_idx = resolve_column(ws, smiles_column, header_row=header_row)
    image_col_idx = resolve_column(ws, image_column, header_row=header_row) if (
        isinstance(image_column, str)
        and image_column.strip().isdigit()
    ) else excel_col_to_index(image_column)

    if start_row is None:
        start_row = header_row + 1

    ws.cell(row=header_row, column=image_col_idx, value=image_header)
    autoset_column_width(ws, image_col_idx, 32)

    temp_dir = tempfile.mkdtemp(prefix="smiles_structures_")

    try:
        for row_idx in range(start_row, ws.max_row + 1):
            smiles_value = ws.cell(row=row_idx, column=smiles_col_idx).value

            if smiles_value is None or str(smiles_value).strip() == "":
                continue

            mol = smiles_to_mol(smiles_value)

            if mol is None:
                ws.cell(row=row_idx, column=image_col_idx, value="Invalid SMILES")
                continue

            png_path = os.path.join(temp_dir, f"mol_row_{row_idx}.png")
            make_structure_png(
                mol,
                png_path,
                img_width=img_width,
                img_height=img_height,
            )

            img = XLImage(png_path)
            anchor = ws.cell(row=row_idx, column=image_col_idx).coordinate
            ws.add_image(img, anchor)

            ws.row_dimensions[row_idx].height = row_height

        wb.save(workbook_path)

    finally:
        try:
            wb.close()
        except Exception:
            pass
        shutil.rmtree(temp_dir, ignore_errors=True)


def parse_args():
    parser = argparse.ArgumentParser(
        description="Insert chemical structure images into Excel from a SMILES column."
    )
    parser.add_argument("--input", required=True, help="Input .csv or .xlsx file")
    parser.add_argument("--output", required=True, help="Output .xlsx file")
    parser.add_argument("--sheet-name", default=None, help="Worksheet name")
    parser.add_argument(
        "--smiles-column",
        required=True,
        help="SMILES column: header name, Excel letter, or 1-based index",
    )
    parser.add_argument(
        "--image-column",
        default="B",
        help="Image column: Excel letter or 1-based index",
    )
    parser.add_argument("--header-row", type=int, default=1, help="Header row")
    parser.add_argument("--start-row", type=int, default=None, help="First data row")
    parser.add_argument("--image-header", default="Structure", help="Image column header")
    parser.add_argument("--img-width", type=int, default=250, help="Image width")
    parser.add_argument("--img-height", type=int, default=180, help="Image height")
    parser.add_argument("--row-height", type=float, default=140, help="Excel row height")
    return parser.parse_args()


def main():
    args = parse_args()

    input_path = Path(args.input)
    output_path = Path(args.output)

    if output_path.suffix.lower() != ".xlsx":
        print("Error: output file must be .xlsx", file=sys.stderr)
        sys.exit(1)

    if not input_path.exists():
        print(f"Error: input file not found: {input_path}", file=sys.stderr)
        sys.exit(1)

    try:
        workbook_path = prepare_workbook_from_input(
            input_path=input_path,
            output_path=output_path,
            sheet_name=args.sheet_name,
        )

        process_workbook(
            workbook_path=workbook_path,
            sheet_name=args.sheet_name,
            smiles_column=args.smiles_column,
            image_column=args.image_column,
            header_row=args.header_row,
            start_row=args.start_row,
            image_header=args.image_header,
            img_width=args.img_width,
            img_height=args.img_height,
            row_height=args.row_height,
        )

        print(f"Done. Saved output to: {workbook_path}")

    except Exception as e:
        print(f"Error: {e}", file=sys.stderr)
        sys.exit(1)


if __name__ == "__main__":
    main()